from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook
import datetime
from tabulate import tabulate
from openpyxl.utils.dataframe import dataframe_to_rows


# dictionary that has Facility Name : [website, sheet name, sheet name]
FACILITY_DICT = {"Nicholas Recreation Center": ["https://recwell.wisc.edu/locations/nick/", "Nicholas Raw", "Nicholas Refined"],
                 "Bakke Recreation & Wellbeing Center": ["https://recwell.wisc.edu/bakke/", "Bakke Raw", "Bakke Refined"]
                 }
response = requests.get(
    'https://goboardapi.azurewebsites.net/api/FacilityCount/GetCountsByAccount?AccountAPIKey=7938FC89-A15C-492D-9566-12C961BC1F27')
df = pd.DataFrame(response.json())


def convertTime(time_str):
    hour = pd.to_datetime(time_str).hour
    if (hour < 12):
        time = str(hour) + ":00 AM"
    elif (hour == 12):
        time = str(hour) + ":00 PM"
    else:
        time = str(hour - 12) + ":00 PM"
    return time


def checkEndOfWeek():
    """Check if the current day is Friday and the current time is after 11:00 PM (to trigger cleanup() method)"""
    now = datetime.datetime.now()
    if now.weekday() == 5 and now.hour >= 23:
        return True
    else:
        return False


def loadData(facility, df):
    """add data from live tracker to excel"""
    # process dataframe
    facility_df = df[df['FacilityName'].str.strip() == facility]
    facility_df = facility_df[['LocationName',
                               'LastUpdatedDateAndTime', 'LastCount', 'IsClosed']]
    facility_df['Time'] = facility_df['LastUpdatedDateAndTime'].apply(
        convertTime)
    facility_df['Date'] = facility_df['LastUpdatedDateAndTime'].apply(
        lambda x: datetime.datetime.strptime(x[:18], "%Y-%m-%dT%H:%M:%S").date())
    facility_df['LastCount'] = df['LastCount'].astype(int)
    facility_df['LastCount'] = facility_df.apply(
        lambda row: None if row['IsClosed'] else row['LastCount'], axis=1)
    facility_df = facility_df[['LocationName', 'Date', 'Time', 'LastCount']]

    # load raw data
    workbook = load_workbook("Live Usage Data.xlsx")
    sheet = workbook[FACILITY_DICT[facility][1]]
    rows = facility_df.values.tolist()
    for row in rows:
        sheet.append(row)
    workbook.save("Live Usage Data.xlsx")
    cleanup(facility)
    refineData(facility, workbook)

def cleanup(facility):
    """remove duplicate rows in Live Usage Data excel 'Raw' sheets so it doesn't get cluttered. Potential idea: delete data that has been there for over 2 months"""
    clean_df = pd.read_excel("Live Usage Data.xlsx",
                           sheet_name=FACILITY_DICT[facility][1])
    clean_df = clean_df.drop_duplicates().reset_index(drop=True)
    clean_df['Date'] = clean_df['Date'].apply(lambda x: datetime.datetime.strptime(str(x)[:10], "%Y-%m-%d").date())
    workbook = load_workbook("Live Usage Data.xlsx")
    sheet = workbook[FACILITY_DICT[facility][1]]
    sheet.delete_rows(1, sheet.max_row)
    for row in dataframe_to_rows(clean_df, index=False, header=True):
        sheet.append(row)
    workbook.save("Live Usage Data.xlsx")
    workbook.close()

def refineData(facility, workbook):
    """Pivot table to match desired formatting"""
    raw_df = pd.read_excel("Live Usage Data.xlsx", sheet_name=FACILITY_DICT[facility][1])
    raw_df = raw_df.drop_duplicates(subset=['Location', 'Date', 'Time']).reset_index(drop=True)
    pivot_df = raw_df.pivot(
        index=['Location', 'Date'], columns='Time', values='Count').reset_index()
    pivot_df['Date'] = pivot_df['Date'].apply(lambda x: datetime.datetime.strptime(str(x)[:10], "%Y-%m-%d").date())
    sheet = workbook[FACILITY_DICT[facility][2]]
    sheet.delete_rows(1, sheet.max_row)
    for row in dataframe_to_rows(pivot_df, index=False, header=True):
        sheet.append(row)
    workbook.save("Live Usage Data.xlsx")
    workbook.close()


def main():
    for facility in FACILITY_DICT:
        loadData(facility, df)
        if (checkEndOfWeek()):
            cleanup(facility)
        cleanup(facility)

main()
